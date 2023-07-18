import datetime
import random

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROW_START_DATA = 4
COL_ROW_ID = 1
COL_COMPANY = 2
COL_FACT_QLIQ_DATA1 = 3
COL_FACT_QLIQ_DATA2 = 4
COL_FACT_QOIL_DATA1 = 5
COL_FACT_QOIL_DATA2 = 6
COL_FORECAST_QLIQ_DATA1 = 7
COL_FORECAST_QLIQ_DATA2 = 8
COL_FORECAST_QOIL_DATA1 = 9
COL_FORECAST_QOIL_DATA2 = 10
START_DATE = datetime.date(2023, 1, 1)
DATE_RANDOM_DELTA = 4


def open_sheet(path) -> Worksheet:
    wb = load_workbook(path)
    return wb.active


def parse_sheet(sheet: Worksheet) -> list[dict]:
    rows = []
    for i in range(ROW_START_DATA, sheet.max_row + 1):
        row = {
            'row_id': sheet.cell(row=i, column=COL_ROW_ID).value,
            'company': sheet.cell(row=i, column=COL_COMPANY).value,
            'fact_qliq_data1': sheet.cell(
                row=i, column=COL_FACT_QLIQ_DATA1).value,
            'fact_qliq_data2': sheet.cell(
                row=i, column=COL_FACT_QLIQ_DATA2).value,
            'fact_qoil_data1': sheet.cell(
                row=i, column=COL_FACT_QOIL_DATA1).value,
            'fact_qoil_data2': sheet.cell(
                row=i, column=COL_FACT_QOIL_DATA2).value,
            'forecast_qliq_data1': sheet.cell(
                row=i, column=COL_FORECAST_QLIQ_DATA1).value,
            'forecast_qliq_data2': sheet.cell(
                row=i, column=COL_FORECAST_QLIQ_DATA2).value,
            'forecast_qoil_data1': sheet.cell(
                row=i, column=COL_FORECAST_QOIL_DATA1).value,
            'forecast_qoil_data2': sheet.cell(
                row=i, column=COL_FORECAST_QOIL_DATA2).value
        }
        rows.append(row)
    return rows


def _add_date_by_company(rows: list[dict]):
    # В задании не указан алгоритм добавления даты. Здесь предполагается, что
    # в исходных данных указаны суммарные значения за день по каждой компании.
    date = START_DATE
    companies_per_day = []
    for row in rows:
        if row['company'] in companies_per_day:
            companies_per_day.clear()
            date += datetime.timedelta(days=1)
        companies_per_day.append(row['company'])
        row['on_date'] = date


def _add_date_by_random(rows: list[dict]):
    for row in rows:
        delta = random.randrange(0, DATE_RANDOM_DELTA)
        date = START_DATE + datetime.timedelta(days=delta)
        row['on_date'] = date


def add_date_column(rows: list[dict]):
    if hasattr(settings, 'DATE_COL_ALGO'):
        if settings.DATE_COL_ALGO == 'random':
            _add_date_by_random(rows)
        elif settings.DATE_COL_ALGO == 'company':
            _add_date_by_company(rows)
        else:
            raise ValueError(
                f'Unknown algo: {settings.DATE_COL_ALGO}. '
                'Please correct your settings.')
    else:
        _add_date_by_company(rows)
